"""Generate .xlsx files from Markdown table content using openpyxl."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .._md_lex import MdToken, lex_markdown


_MAX_SHEETS = 20
_MAX_ROWS_PER_SHEET = 1000

# Styles
_THOUSANDS_SEP_RE = re.compile(r"^\d{1,3}(,\d{3})+$")
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_HEADER_FONT = Font(name="宋体", size=10, bold=True, color="FFFFFF")
_DATA_FONT = Font(name="宋体", size=10)
_SUM_FONT = Font(name="宋体", size=10, bold=True)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_DOUBLE_TOP_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="double"),
    bottom=Side(style="thin"),
)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_RIGHT = Alignment(horizontal="right", vertical="center")


def write_xlsx_from_sheets(
    path: Path,
    sheets_data: list[dict[str, str]],
) -> str:
    """Create an .xlsx file from a list of sheet definitions.

    Args:
        path: Output file path.
        sheets_data: List of {"name": str, "content": str} where content is Markdown.

    Returns:
        A human-readable summary string.
    """
    if len(sheets_data) > _MAX_SHEETS:
        raise ValueError(f"Sheet 数量超限: {len(sheets_data)} (上限 {_MAX_SHEETS})")

    wb = Workbook()
    # Remove the default sheet — we'll create our own
    wb.remove(wb.active)

    for sheet_def in sheets_data:
        sheet_name = sheet_def.get("name", "Sheet")
        content = sheet_def.get("content", "")
        _write_sheet(wb, sheet_name, content)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))

    size_kb = path.stat().st_size / 1024
    summary_parts = [f"文件已生成: {path} ({size_kb:.1f} KB)"]

    detail_parts: list[str] = []
    for ws in wb.worksheets:
        row_count = ws.max_row - 1 if ws.max_row and ws.max_row > 1 else 0
        detail_parts.append(f"Sheet \"{ws.title}\" ({row_count} 行)")

    if detail_parts:
        summary_parts.append("包含: " + "、".join(detail_parts))

    return "\n".join(summary_parts)


def _write_sheet(wb: Workbook, name: str, content: str) -> None:
    """Write a single worksheet from Markdown content."""
    ws = wb.create_sheet(title=name[:31])  # Excel sheet name max 31 chars
    tokens = lex_markdown(content)

    current_row = 1
    table_count = 0

    for token in tokens:
        if token.type == "table":
            table_count += 1
            current_row = _write_table(ws, token, current_row)
            current_row += 1  # blank row between tables
        elif token.type == "heading":
            ws.cell(row=current_row, column=1, value=token.payload["text"])
            ws.cell(row=current_row, column=1).font = Font(name="宋体", size=12, bold=True)
            current_row += 1
        elif token.type == "paragraph":
            ws.cell(row=current_row, column=1, value=token.payload.get("text", ""))
            ws.cell(row=current_row, column=1).font = _DATA_FONT
            current_row += 1


def _write_table(ws: Any, token: MdToken, start_row: int) -> int:
    """Write a table token into the worksheet, return next available row."""
    headers = token.payload.get("headers", [])
    rows = token.payload.get("rows", [])
    col_count = len(headers)

    if col_count == 0:
        return start_row

    # Enforce row limit
    if len(rows) > _MAX_ROWS_PER_SHEET:
        rows = rows[:_MAX_ROWS_PER_SHEET]

    # Detect which columns look numeric (for alignment)
    numeric_cols = _detect_numeric_columns(rows)

    # Header row
    row_idx = start_row
    for ci, h in enumerate(headers):
        cell = ws.cell(row=row_idx, column=ci + 1, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER
    row_idx += 1

    # Data rows
    for ri, row in enumerate(rows):
        is_last = ri == len(rows) - 1
        is_sum_row = _is_summary_row(row, headers)

        for ci in range(col_count):
            raw = row[ci] if ci < len(row) else ""
            cell = ws.cell(row=row_idx, column=ci + 1)
            _set_cell_value(cell, raw)

            if is_sum_row:
                cell.font = _SUM_FONT
                cell.border = _DOUBLE_TOP_BORDER
            else:
                cell.font = _DATA_FONT
                cell.border = _THIN_BORDER

            # Alignment
            if ci in numeric_cols and not isinstance(cell.value, str):
                cell.alignment = _RIGHT
            else:
                cell.alignment = _CENTER

        row_idx += 1

    # Auto-fit column widths
    for ci in range(col_count):
        max_len = len(str(headers[ci]))
        for row in rows:
            val = row[ci] if ci < len(row) else ""
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(ci + 1)].width = max(max_len * 1.2, 8)

    return row_idx


def _detect_numeric_columns(rows: list[list[str]]) -> set[int]:
    """Detect columns that contain mostly numeric values."""
    if not rows:
        return set()

    numeric_cols: set[int] = set()
    col_count = max(len(r) for r in rows) if rows else 0

    for ci in range(col_count):
        numeric_count = 0
        total = 0
        for row in rows:
            if ci < len(row):
                val = row[ci].strip()
                if val and not val.startswith("="):
                    total += 1
                    if _is_number(val):
                        numeric_count += 1
        if total > 0 and numeric_count / total > 0.5:
            numeric_cols.add(ci)

    return numeric_cols


def _is_number(text: str) -> bool:
    """Check if text represents a number (including with thousands separator)."""
    cleaned = text.replace(",", "")
    try:
        float(cleaned)
        return True
    except ValueError:
        return False


def _is_summary_row(row: list[str], headers: list[str]) -> bool:
    """Detect if a row is a summary/total row."""
    for val in row:
        v = val.strip()
        if v in ("合计", "总计", "小计", "合计(元)", "合计（元）"):
            return True
    return False


def _set_cell_value(cell: Any, raw: str) -> None:
    """Set a cell's value with type inference: formula, number, or text."""
    text = raw.strip()
    if not text:
        cell.value = ""
        return

    # Formula
    if text.startswith("="):
        cell.value = text
        return

    # Pure integer
    try:
        cell.value = int(text)
        cell.number_format = "#,##0"
        return
    except ValueError:
        pass

    # Pure float
    try:
        cell.value = float(text)
        # Check if it looks like a currency value (has .00 or .XX)
        if "." in text:
            cell.number_format = "#,##0.00"
        else:
            cell.number_format = "#,##0"
        return
    except ValueError:
        pass

    # Thousands-separated number: "1,200" or "1,200.50"
    cleaned = text.replace(",", "")
    if _THOUSANDS_SEP_RE.match(text) or ("," in text and _is_number(text)):
        try:
            val = float(cleaned)
            if val == int(val):
                cell.value = int(val)
            else:
                cell.value = val
            if "." in text:
                cell.number_format = "#,##0.00"
            else:
                cell.number_format = "#,##0"
            return
        except ValueError:
            pass

    # Text
    cell.value = text
