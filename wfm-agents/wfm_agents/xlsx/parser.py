"""Parse .xlsx files into structured dicts using openpyxl."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


_MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB
_MAX_SHEETS = 20
_MAX_ROWS_PER_SHEET = 200


def parse_xlsx(
    path: Path,
    *,
    sheet: str | None = None,
    max_rows_per_sheet: int = _MAX_ROWS_PER_SHEET,
) -> dict[str, Any]:
    """Parse an .xlsx file into a structured dict.

    Returns:
        {
            "metadata": {"title": str, "author": str, "created": str, "sheet_names": [str]},
            "sheets": [{
                "name": str, "col_count": int, "row_count": int,
                "headers": [str], "rows": [[Cell]], "merged_ranges": [str],
                "has_formulas": bool, "truncated": bool,
            }],
            "stats": {"sheets_total": int, "sheets_kept": int, "total_rows": int},
        }
    """
    if path.stat().st_size > _MAX_FILE_SIZE:
        raise ValueError(
            f"文件过大: {path.stat().st_size / 1024 / 1024:.1f} MB (上限 10 MB)"
        )

    wb = load_workbook(str(path), data_only=False)

    props = wb.properties
    metadata = {
        "title": props.title or "",
        "author": props.creator or "",
        "created": props.created.isoformat() if props.created else "",
        "sheet_names": wb.sheetnames,
    }

    all_sheet_names = wb.sheetnames
    target_names = [sheet] if sheet else all_sheet_names
    if sheet and sheet not in all_sheet_names:
        wb.close()
        raise ValueError(f"Sheet 不存在: {sheet}，可用: {', '.join(all_sheet_names)}")

    sheets_data: list[dict[str, Any]] = []
    total_rows = 0

    for name in target_names[:_MAX_SHEETS]:
        ws = wb[name]

        # Collect cell values directly via iter_rows (0-indexed in our array)
        raw_rows: list[list[Any]] = []
        for ri, row in enumerate(ws.iter_rows()):
            if ri >= max_rows_per_sheet + 1:
                break
            raw_rows.append(list(row))

        truncated = ws.max_row is not None and ws.max_row > max_rows_per_sheet + 1

        # Merged ranges
        merged_ranges = [str(mr) for mr in ws.merged_cells.ranges]
        merge_map = _build_merge_map(ws.merged_cells.ranges)

        headers: list[str] = []
        data_rows: list[list[dict[str, Any]]] = []
        has_formulas = False

        for ri, raw_row in enumerate(raw_rows):
            cells: list[dict[str, Any]] = []
            for ci, cell in enumerate(raw_row):
                # For merged cells, look up the anchor cell's value
                if (ri, ci) in merge_map:
                    anchor = merge_map[(ri, ci)]
                    cell = raw_rows[anchor[0]][anchor[1]]

                cell_info = _parse_cell(cell)
                if cell_info.get("formula"):
                    has_formulas = True
                cells.append(cell_info)

            if ri == 0:
                headers = [c["formatted"] or c["value"] or "" for c in cells]
            else:
                data_rows.append(cells)

        col_count = len(headers) if headers else 0
        row_count = len(data_rows)
        total_rows += row_count

        sheets_data.append({
            "name": name,
            "col_count": col_count,
            "row_count": row_count,
            "headers": headers,
            "rows": data_rows,
            "merged_ranges": merged_ranges,
            "has_formulas": has_formulas,
            "truncated": truncated,
        })

    wb.close()

    return {
        "metadata": metadata,
        "sheets": sheets_data,
        "stats": {
            "sheets_total": len(all_sheet_names),
            "sheets_kept": len(sheets_data),
            "total_rows": total_rows,
        },
    }


def _build_merge_map(merged_ranges: Any) -> dict[tuple[int, int], tuple[int, int]]:
    """Build a map from (row, col) of merged cells to their anchor (top-left)."""
    merge_map: dict[tuple[int, int], tuple[int, int]] = {}
    for mr in merged_ranges:
        min_row = mr.min_row - 1  # 0-indexed
        min_col = mr.min_col - 1
        for r in range(mr.min_row - 1, mr.max_row):
            for c in range(mr.min_col - 1, mr.max_col):
                if (r, c) != (min_row, min_col):
                    merge_map[(r, c)] = (min_row, min_col)
    return merge_map


def _parse_cell(cell: Any) -> dict[str, Any]:
    """Parse a single openpyxl cell into a dict."""
    value = cell.value
    formula = None

    if isinstance(value, str) and value.startswith("="):
        formula = value
        # In read_only mode with data_only=False, formula cells return the formula string.
        # We don't have the cached value, so set formatted to the formula.
        return {
            "value": value,
            "formatted": value,
            "type": "formula",
            "formula": formula,
        }

    cell_type = "text"
    formatted = ""

    if value is None:
        cell_type = "text"
        formatted = ""
    elif isinstance(value, bool):
        cell_type = "text"
        formatted = "TRUE" if value else "FALSE"
    elif isinstance(value, (int, float)):
        cell_type = "number"
        formatted = _format_number(value, cell.number_format)
    elif isinstance(value, str):
        cell_type = "text"
        formatted = value
    else:
        cell_type = "text"
        formatted = str(value)

    return {
        "value": value,
        "formatted": formatted,
        "type": cell_type,
        "formula": None,
    }


def _format_number(value: int | float, number_format: str) -> str:
    """Format a numeric value respecting the cell's number format."""
    if not number_format or number_format == "General":
        if isinstance(value, float) and value == int(value):
            return str(int(value))
        return str(value)

    # Apply thousands separator if format contains ','
    if "," in number_format:
        if isinstance(value, float):
            return f"{value:,.2f}"
        return f"{value:,}"

    # Currency
    if "¥" in number_format or "$" in number_format or "￥" in number_format:
        if isinstance(value, float):
            return f"¥{value:,.2f}"
        return f"¥{value:,}"

    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value)


# ── Markdown output ────────────────────────────────────────────────────


def format_xlsx_content(content: dict[str, Any]) -> str:
    """Format parsed xlsx content into Markdown text for LLM consumption."""
    parts: list[str] = []

    meta = content.get("metadata", {})
    title = meta.get("title", "")
    if title:
        parts.append(f"# {title}\n")

    for sheet in content.get("sheets", []):
        name = sheet.get("name", "")
        parts.append(f"## Sheet: {name}\n")

        headers = sheet.get("headers", [])
        rows = sheet.get("rows", [])

        if headers:
            parts.append("| " + " | ".join(headers) + " |")
            parts.append("| " + " | ".join("---" for _ in headers) + " |")

            for row in rows:
                padded = []
                for ci, cell in enumerate(row):
                    if ci < len(headers):
                        formatted = cell.get("formatted", "")
                        if cell.get("formula"):
                            formatted = f"`{cell['formula']}`"
                        padded.append(formatted)
                # Pad to header length
                while len(padded) < len(headers):
                    padded.append("")
                parts.append("| " + " | ".join(padded) + " |")

        # Annotations
        annotations: list[str] = []
        merged = sheet.get("merged_ranges", [])
        if merged:
            annotations.append(f"合并单元格: {', '.join(merged[:5])}")
            if len(merged) > 5:
                annotations[-1] += f" 等 {len(merged)} 处"
        if sheet.get("has_formulas"):
            formula_count = sum(
                1 for row in rows for cell in row if cell.get("formula")
            )
            annotations.append(f"包含公式: 是（{formula_count} 处）")
        if sheet.get("truncated"):
            annotations.append("数据已截断，含更多行")

        if annotations:
            for ann in annotations:
                parts.append(f"> {ann}")

        parts.append("")  # blank line after sheet
        parts.append("---")
        parts.append("")

    stats = content.get("stats", {})
    sheets_total = stats.get("sheets_total", 0)
    total_rows = stats.get("total_rows", 0)
    parts.append(f"> 共 {sheets_total} 个 Sheet，{total_rows} 行数据。")

    return "\n".join(parts)
